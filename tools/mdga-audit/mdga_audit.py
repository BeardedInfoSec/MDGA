"""
MDGA Audit Tool
================================================
On-demand audit for the MDGA federation. Walks an officer through
capturing one-or-more guild rosters from the in-game addon, ingests
a Discord roster CSV, optionally pulls the live website roster,
then prints a categorized punch list and writes a CSV report.

Pure stdlib (no pip install needed). Built into a single .exe via
PyInstaller for non-technical officers.

Run:  python mdga_audit.py
Build: pyinstaller --onefile --name mdga-audit mdga_audit.py
"""
from __future__ import annotations

import csv
import json
import os
import re
import ssl
import subprocess
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

# Bundled into the .exe via PyInstaller. Tested with openpyxl 3.1.x.
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

VERSION = "1.4.0"
DEFAULT_SERVER = "https://mdga.gg"

# Config file lives next to the exe (or the .py during dev). Stores WoW path,
# account, server URL, token, and Discord-CSV folder so subsequent runs can
# skip every prompt by pressing Enter. Pass --reconfigure to rerun the wizard.
def _app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(_app_dir(), "mdga-audit-config.json")
DISCORD_INPUT_DIR = os.path.join(_app_dir(), "discord-reports")
ARCHIVE_DIR = os.path.join(_app_dir(), "archive")


def setup_workdir() -> None:
    """Ensure the input/output directory structure exists. Idempotent."""
    os.makedirs(DISCORD_INPUT_DIR, exist_ok=True)
    os.makedirs(ARCHIVE_DIR, exist_ok=True)


def make_run_dir() -> str:
    """Create archive/<YYYY-MM-DD_HHMMSS>/ for this run and return its path."""
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    run_dir = os.path.join(ARCHIVE_DIR, stamp)
    os.makedirs(os.path.join(run_dir, "wow"), exist_ok=True)
    os.makedirs(os.path.join(run_dir, "discord"), exist_ok=True)
    return run_dir


def archive_snapshot_json(snap: 'GuildSnapshot', dest_dir: str) -> str:
    """Serialize a GuildSnapshot to JSON for re-audit / longitudinal comparison."""
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", snap.guild_name).strip("_") or "guild"
    realm = (snap.realm_slug or "unknown").replace("/", "_")
    path = os.path.join(dest_dir, f"snapshot_{safe}_{realm}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump({
            "guild_name": snap.guild_name,
            "realm_slug": snap.realm_slug,
            "captured_at": snap.captured_at,
            "members": snap.members,
        }, f, indent=2, default=str)
    return path


def load_config() -> dict[str, Any]:
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_config(config: dict[str, Any]) -> None:
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
    except Exception as e:
        print(f"  ! Couldn't save config: {e}")

# Force UTF-8 on stdout so the pretty arrows / checkmarks below don't crash on
# Windows consoles whose default codepage is cp1252. No-op on Python builds
# that don't support reconfigure() or platforms that don't need it.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# ──────────────────────────────────────────────────────────────────
# Lua parser (hand-rolled, no third-party dependency)
# Recursive-descent. No eval. Handles tables, strings, numbers,
# booleans, nil, line + block comments. Sufficient for WoW
# SavedVariables (which is a constrained subset of Lua).
# ──────────────────────────────────────────────────────────────────


class LuaParseError(Exception):
    pass


class LuaParser:
    def __init__(self, source: str):
        self.src = source
        self.pos = 0
        self.n = len(source)

    def _peek(self) -> str:
        return self.src[self.pos] if self.pos < self.n else ""

    def _advance(self) -> str:
        ch = self.src[self.pos]
        self.pos += 1
        return ch

    def _expect(self, ch: str) -> None:
        if self._peek() != ch:
            raise LuaParseError(f"Expected '{ch}' at pos {self.pos}, got '{self._peek()}'")
        self._advance()

    def _skip_ws(self) -> None:
        while self.pos < self.n:
            ch = self.src[self.pos]
            if ch in " \t\r\n":
                self.pos += 1
                continue
            # Comments: -- line, --[[ block ]]
            if ch == "-" and self.pos + 1 < self.n and self.src[self.pos + 1] == "-":
                self.pos += 2
                if self.pos + 1 < self.n and self.src[self.pos] == "[" and self.src[self.pos + 1] == "[":
                    self.pos += 2
                    while self.pos + 1 < self.n:
                        if self.src[self.pos] == "]" and self.src[self.pos + 1] == "]":
                            self.pos += 2
                            break
                        self.pos += 1
                else:
                    while self.pos < self.n and self.src[self.pos] != "\n":
                        self.pos += 1
                continue
            break

    def parse_file(self) -> dict[str, Any]:
        """Parse a SavedVariables file into a dict of top-level globals."""
        result: dict[str, Any] = {}
        self._skip_ws()
        while self.pos < self.n:
            name = self._read_identifier()
            self._skip_ws()
            self._expect("=")
            self._skip_ws()
            result[name] = self._read_value()
            self._skip_ws()
        return result

    def _read_identifier(self) -> str:
        start = self.pos
        while self.pos < self.n and (self.src[self.pos].isalnum() or self.src[self.pos] == "_"):
            self.pos += 1
        if start == self.pos:
            raise LuaParseError(f"Expected identifier at pos {self.pos}")
        return self.src[start:self.pos]

    def _read_value(self) -> Any:
        self._skip_ws()
        ch = self._peek()
        if ch == "{":
            return self._read_table()
        if ch in '"\'':
            return self._read_string(ch)
        if ch == "[":
            # Long-bracket string [[...]] or [=[...]=]
            return self._read_long_string()
        if ch == "-" or ch.isdigit():
            return self._read_number()
        # keyword: true / false / nil
        kw = self._read_identifier()
        if kw == "true":
            return True
        if kw == "false":
            return False
        if kw == "nil":
            return None
        raise LuaParseError(f"Unexpected keyword '{kw}' at pos {self.pos}")

    def _read_string(self, quote: str) -> str:
        self._expect(quote)
        out: list[str] = []
        while self.pos < self.n:
            ch = self._advance()
            if ch == quote:
                return "".join(out)
            if ch == "\\" and self.pos < self.n:
                esc = self._advance()
                out.append({"n": "\n", "r": "\r", "t": "\t",
                            "\\": "\\", "'": "'", '"': '"',
                            "0": "\0"}.get(esc, esc))
            else:
                out.append(ch)
        raise LuaParseError("Unterminated string")

    def _read_long_string(self) -> str:
        # [[...]] or [=...=[...]=...=]
        self._expect("[")
        eq_count = 0
        while self._peek() == "=":
            self._advance()
            eq_count += 1
        self._expect("[")
        closer = "]" + ("=" * eq_count) + "]"
        end = self.src.find(closer, self.pos)
        if end < 0:
            raise LuaParseError("Unterminated long string")
        s = self.src[self.pos:end]
        self.pos = end + len(closer)
        # Lua long strings drop a leading newline if present
        if s.startswith("\n"):
            s = s[1:]
        return s

    def _read_number(self) -> int | float:
        start = self.pos
        if self._peek() == "-":
            self._advance()
        # Hex
        if self.src[self.pos:self.pos + 2] in ("0x", "0X"):
            self.pos += 2
            while self.pos < self.n and self.src[self.pos] in "0123456789abcdefABCDEF":
                self.pos += 1
            return int(self.src[start:self.pos], 16)
        while self.pos < self.n and (self.src[self.pos].isdigit() or self.src[self.pos] in ".eE+-"):
            ch = self.src[self.pos]
            # don't gobble trailing - that's a separator
            if ch in "+-" and self.pos > start and self.src[self.pos - 1] not in "eE":
                break
            self.pos += 1
        text = self.src[start:self.pos]
        try:
            if "." in text or "e" in text or "E" in text:
                return float(text)
            return int(text)
        except ValueError:
            raise LuaParseError(f"Invalid number '{text}' at pos {start}")

    def _read_table(self) -> dict[Any, Any] | list[Any]:
        self._expect("{")
        self._skip_ws()
        result_dict: dict[Any, Any] = {}
        next_array_idx = 1
        is_array = True
        while self._peek() != "}":
            self._skip_ws()
            key: Any = None
            # ["key"] = value  OR  [123] = value  OR  identifier = value
            if self._peek() == "[":
                self._advance()
                self._skip_ws()
                key = self._read_value()
                self._skip_ws()
                self._expect("]")
                self._skip_ws()
                self._expect("=")
                self._skip_ws()
                is_array = False
            else:
                # try identifier = value (key lookup), else positional value
                save = self.pos
                # peek for identifier
                if self.src[self.pos].isalpha() or self.src[self.pos] == "_":
                    ident_start = self.pos
                    while self.pos < self.n and (self.src[self.pos].isalnum() or self.src[self.pos] == "_"):
                        self.pos += 1
                    ident_end = self.pos
                    self._skip_ws()
                    if self._peek() == "=":
                        key = self.src[ident_start:ident_end]
                        self._advance()  # consume =
                        self._skip_ws()
                        is_array = False
                    else:
                        # not an assignment; rewind and treat as value
                        self.pos = save
                        key = next_array_idx
                        next_array_idx += 1
                else:
                    key = next_array_idx
                    next_array_idx += 1
            value = self._read_value()
            result_dict[key] = value
            self._skip_ws()
            if self._peek() in ",;":
                self._advance()
                self._skip_ws()
        self._expect("}")
        # If purely positional 1..N, return list (more Pythonic)
        if is_array and result_dict and all(k == i + 1 for i, k in enumerate(sorted(result_dict.keys()))):
            return [result_dict[k] for k in sorted(result_dict.keys())]
        return result_dict


def parse_lua_file(path: str) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        src = f.read()
    # Strip BOM if present
    if src.startswith("﻿"):
        src = src[1:]
    return LuaParser(src).parse_file()


# ──────────────────────────────────────────────────────────────────
# WoW install detection
# ──────────────────────────────────────────────────────────────────

DEFAULT_WOW_PATHS = [
    r"C:\Program Files (x86)\World of Warcraft\_retail_",
    r"C:\Program Files\World of Warcraft\_retail_",
    r"D:\World of Warcraft\_retail_",
    r"D:\Program Files (x86)\World of Warcraft\_retail_",
]


def detect_wow_path() -> str | None:
    """Standard install paths first (most common); registry only as a fallback so
    that a stray Battle.net registry entry pointing at a private-server launcher
    (e.g. Ascension) doesn't beat out a perfectly good retail install."""
    for p in DEFAULT_WOW_PATHS:
        if os.path.isdir(p):
            return p
    if sys.platform == "win32":
        for key in [
            r"HKLM\Software\Wow6432Node\Blizzard Entertainment\World of Warcraft",
            r"HKLM\Software\Blizzard Entertainment\World of Warcraft",
        ]:
            try:
                out = subprocess.run(
                    ["reg.exe", "query", key, "/v", "InstallPath"],
                    capture_output=True, text=True, timeout=5
                )
                if out.returncode == 0:
                    m = re.search(r"InstallPath\s+REG_SZ\s+(.+)", out.stdout)
                    if m:
                        base = m.group(1).strip().rstrip("\\")
                        retail = os.path.join(base, "_retail_")
                        # Only trust a registry hit that actually looks like retail WoW.
                        if os.path.isfile(os.path.join(retail, "Wow.exe")):
                            return retail
                        if os.path.isfile(os.path.join(base, "Wow.exe")):
                            return base
            except Exception:
                pass
    # Last resort: hand back the canonical default even if it doesn't exist on
    # this machine, so the wizard prompt shows a sensible suggestion.
    return DEFAULT_WOW_PATHS[0]


def list_account_folders(wow_path: str) -> list[str]:
    acct_dir = os.path.join(wow_path, "WTF", "Account")
    if not os.path.isdir(acct_dir):
        return []
    return sorted(
        d for d in os.listdir(acct_dir)
        if os.path.isdir(os.path.join(acct_dir, d)) and not d.startswith(".")
    )


# ──────────────────────────────────────────────────────────────────
# Multi-guild capture flow
# ──────────────────────────────────────────────────────────────────


@dataclass
class GuildSnapshot:
    """One captured guild roster from a /reload-flushed SavedVariables."""
    guild_name: str
    realm_slug: str | None
    captured_at: int
    members: list[dict[str, Any]] = field(default_factory=list)


def watch_for_capture(sv_path: str, baseline_mtime: float, timeout_sec: int = 600) -> GuildSnapshot | None:
    """Block until SV file is rewritten with a fresh pendingReport, then return it."""
    deadline = time.time() + timeout_sec
    last_print = 0.0
    while time.time() < deadline:
        try:
            mtime = os.path.getmtime(sv_path)
        except FileNotFoundError:
            mtime = 0
        now = time.time()
        if now - last_print > 30:
            remaining = int(deadline - now)
            print(f"  ... waiting for /reload (timeout in {remaining}s)")
            last_print = now

        if mtime > baseline_mtime + 0.5:  # 0.5s grace for filesystem rounding
            # File changed — give chokidar-style settle time then parse
            time.sleep(2.0)
            try:
                data = parse_lua_file(sv_path)
            except (LuaParseError, OSError) as e:
                print(f"  ! Parse failed: {e}. Waiting for next /reload...")
                baseline_mtime = mtime
                continue

            mdga = data.get("MDGA_Data") or {}
            report = mdga.get("pendingReport") if isinstance(mdga, dict) else None
            if not report:
                print("  ! SV updated but no pendingReport found. Did you click 'Generate Report' before /reload?")
                baseline_mtime = mtime
                continue

            roster = report.get("roster") or []
            if not isinstance(roster, list):
                roster = list(roster.values()) if isinstance(roster, dict) else []
            guild_info = report.get("guildInfo") or {}
            guild_name = guild_info.get("name") or "Unknown Guild"
            # realm slug: take from first member
            realm_slug = None
            for m in roster:
                if isinstance(m, dict) and m.get("realmSlug"):
                    realm_slug = m["realmSlug"]
                    break

            return GuildSnapshot(
                guild_name=str(guild_name),
                realm_slug=realm_slug,
                captured_at=int(report.get("generatedAt") or time.time()),
                members=[m for m in roster if isinstance(m, dict)],
            )
        time.sleep(2.0)
    return None


# ──────────────────────────────────────────────────────────────────
# Discord roster CSV (the format produced by your bot in #bot-spam)
# ──────────────────────────────────────────────────────────────────
#
# The bot exports two CSVs (one per main-rank role: Durotarian + Elwynnian)
# with this schema:
#   "User", "ID", "Nickname", "@everyone", <role_1>, <role_2>, ..., <role_N>
# Each role column is empty for users without the role, or contains the role
# name verbatim for users with it. Nickname is "Charname-Realm".
#
# Roles we care about for the audit (in-game rank ↔ Discord role):
MAIN_ROLES = {"Durotarian", "Elwynnian"}                       # full member
SUB_MAIN_ROLES = {"MDGA Friend"}                               # social-only
GAME_RANK_ROLES = {                                            # mirror of in-game ranks
    "Warchief", "Warlord", "War Council", "RWC",
    "Marshal", "Knight", "Guardian", "Champion",
    "Honorbound", "Durotari", "Sapphire",
}
SKIP_ROLE_COLUMNS = {                                          # bot infrastructure / not ranks
    "User", "ID", "Nickname", "@everyone",
    "WoW Blue Post Tracker", "Message Scheduler", "StickyBot",
    "VoiceMaster", "Verification Bot V2", "MemberList",
    "Guilds of WoW", "verifications", "Open Ticket",
    "Raid-Helper", "Pancake", "Fight Club", "RBG Captain", "RBG1",
    "carl-bot", "PvP Coach", "FROST GOD BEOND", "Server Booster",
    "MEGA Jr Officer", "MDGA Jr Officer 3", "MDGA Jr Officer 2",
    "MDGA Jr Officer 1", "Ticket Tool", "Support",
}


@dataclass
class DiscordMember:
    user: str           # Discord username (e.g. "joobies1123")
    user_id: str        # Discord snowflake
    nickname: str       # raw nickname field
    char_name: str      # parsed character name (lowercased)
    realm_slug: str     # parsed realm slug (Blizzard format: "tichondrius", "zuljin", "moon-guard"...)
    roles: set[str]     # all Discord roles this user has

    @property
    def main_role(self) -> str | None:
        for r in MAIN_ROLES:
            if r in self.roles:
                return r
        return None

    @property
    def is_main(self) -> bool:
        return bool(self.main_role)


# Realms whose Blizzard slug doesn't follow the mechanical rules. Keys are the
# normalized output of the algorithm; values are the actual Blizzard slug.
# (Add to this as new mismatches surface in audit reports.)
REALM_ALIASES = {
    "area52": "area-52",
    "wyrmrestaccord": "wyrmrest-accord",
    "forgottencoast": "the-forgotten-coast",
    "blacksgridge": "blackrock",  # placeholder example
}


def normalize_realm(realm_text: str) -> str:
    """Convert Discord nickname realm chunk (`Tichondrius` / `Mal'Ganis` / `MoonGuard` /
    `Lightning'sBlade`) into the Blizzard slug format (`tichondrius` / `malganis` /
    `moon-guard` / `lightnings-blade`).

    Strategy: insert hyphens at camelCase boundaries FIRST (so apostrophes act as
    barriers and don't create spurious hyphens — `Mal'Ganis` stays one word), THEN
    strip apostrophes, then apply the alias table for known oddballs."""
    if not realm_text:
        return ""
    s = realm_text
    # Insert hyphens at camelCase boundaries while apostrophes are still in place.
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", "-", s)
    s = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", "-", s)
    # Now collapse apostrophes (don't hyphenate them).
    s = s.replace("'", "").replace("’", "")
    s = s.lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9-]", "", s)
    return REALM_ALIASES.get(s, s)


def parse_discord_nickname(nick: str) -> tuple[str, str]:
    """`Polychange-Tichondrius` → ('polychange', 'tichondrius').
    Falls back to (raw_lower, '') when the nick doesn't follow the format."""
    if not nick or "-" not in nick:
        return (nick.lower().strip(), "")
    name_part, _, realm_part = nick.rpartition("-")
    return (name_part.strip().lower(), normalize_realm(realm_part.strip()))


def read_discord_csv(path: str) -> list[DiscordMember]:
    members: list[DiscordMember] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            roles: set[str] = set()
            for col, val in row.items():
                if not col or col in SKIP_ROLE_COLUMNS:
                    continue
                if (val or "").strip():
                    roles.add(col.strip())
            nick = (row.get("Nickname") or "").strip()
            char, realm = parse_discord_nickname(nick)
            members.append(DiscordMember(
                user=(row.get("User") or "").strip(),
                user_id=(row.get("ID") or "").strip(),
                nickname=nick,
                char_name=char,
                realm_slug=realm,
                roles=roles,
            ))
    return members


def read_discord_csvs(paths: list[str]) -> list[DiscordMember]:
    """Combine multiple CSV exports (e.g. Durotarian + Elwynnian filters) into one
    de-duplicated list keyed by Discord user ID — same user shows up in both exports
    if they hold both roles, but we want one row per user with both roles merged."""
    by_id: dict[str, DiscordMember] = {}
    for p in paths:
        for m in read_discord_csv(p):
            if not m.user_id:
                continue
            if m.user_id in by_id:
                by_id[m.user_id].roles |= m.roles
            else:
                by_id[m.user_id] = m
    return list(by_id.values())


# ──────────────────────────────────────────────────────────────────
# Website client (optional)
# ──────────────────────────────────────────────────────────────────


# Decode the JWT payload locally (no signature check — just want the username
# for the report header). JWT is `header.payload.signature` with each part
# being URL-safe base64 of JSON.
def decode_jwt_payload(token: str) -> dict[str, Any]:
    import base64
    try:
        parts = token.split(".")
        if len(parts) < 2:
            return {}
        payload = parts[1]
        # Pad to multiple of 4 (URL-safe base64 strips padding)
        payload += "=" * (-len(payload) % 4)
        return json.loads(base64.urlsafe_b64decode(payload).decode("utf-8"))
    except Exception:
        return {}


def fetch_json(url: str, token: str, timeout: int = 30) -> Any:
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "User-Agent": f"mdga-audit/{VERSION}",
    })
    ctx = ssl.create_default_context()
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.load(resp)


def fetch_website_audit_data(server_url: str, token: str) -> dict[str, Any]:
    """Pull the existing reconciliation reports from mdga.gg for cross-checks."""
    base = server_url.rstrip("/")
    out: dict[str, Any] = {"guild_gaps": [], "discord_orphans": [], "spelling_mismatches": [], "errors": []}
    for slug, key in [
        ("guild-gaps?export_all=true", "guild_gaps"),
        ("discord-orphans?export_all=true", "discord_orphans"),
        ("spelling-mismatches?export_all=true", "spelling_mismatches"),
    ]:
        try:
            data = fetch_json(f"{base}/api/reports/{slug}", token)
            # endpoints return varying shapes — store raw for now
            if isinstance(data, dict):
                out[key] = data.get("rows") or data.get("results") or data.get("data") or []
            elif isinstance(data, list):
                out[key] = data
        except urllib.error.HTTPError as e:
            out["errors"].append(f"{key}: HTTP {e.code} — {e.reason}")
        except Exception as e:
            out["errors"].append(f"{key}: {e}")
    return out


# ──────────────────────────────────────────────────────────────────
# Audit logic
# ──────────────────────────────────────────────────────────────────

ALT_NOTE_RE = re.compile(r"^(?P<main>[A-Za-z'\-]+)['’]s\s*alt\s*$", re.IGNORECASE)
MAIN_RANK_NAMES = {"DUROTARIAN", "ELWYNNIAN"}


@dataclass
class Finding:
    category: str
    character: str
    realm: str
    detail: str
    severity: str = "info"  # info / warn / action

    def as_row(self) -> dict[str, str]:
        return {
            "Category": self.category,
            "Severity": self.severity,
            "Character": self.character,
            "Realm": self.realm,
            "Detail": self.detail,
        }


def audit(
    snapshots: list[GuildSnapshot],
    discord_members: list[DiscordMember] | None,
    website: dict[str, Any] | None,
) -> list[Finding]:
    findings: list[Finding] = []

    # Flatten in-game roster across all captured guilds.
    in_game: list[dict[str, Any]] = []
    main_index: dict[str, dict[str, Any]] = {}  # name lower → main char dict
    alt_pointers: list[tuple[dict[str, Any], str]] = []  # (alt_char, claimed_main_name_lower)
    by_char_realm: dict[tuple[str, str], dict[str, Any]] = {}
    # Realms we captured a roster for — the only realms where we can authoritatively
    # say "this character is/isn't in our guild." For Discord members on uncaptured
    # realms, all we can honestly report is "we didn't scan that guild."
    captured_realms: set[str] = set()
    for snap in snapshots:
        for m in snap.members:
            entry = dict(m)
            entry["_guild"] = snap.guild_name
            in_game.append(entry)
            char = (m.get("name") or "").lower()
            realm = (m.get("realmSlug") or "").lower()
            if realm:
                captured_realms.add(realm)
            by_char_realm[(char, realm)] = entry
            rank_name = (m.get("rankName") or "").upper().strip()
            if rank_name in MAIN_RANK_NAMES:
                main_index[char] = entry
            mm = ALT_NOTE_RE.match(m.get("officerNote") or "")
            if mm:
                alt_pointers.append((entry, mm.group("main").lower()))
        if snap.realm_slug:
            captured_realms.add(snap.realm_slug.lower())

    # CHECK 1: Mains who left in-game but their alts remain.
    orphaned_mains: dict[str, list[dict[str, Any]]] = {}
    for alt, main_name in alt_pointers:
        if main_name not in main_index:
            orphaned_mains.setdefault(main_name, []).append(alt)
    for main_name, alts in sorted(orphaned_mains.items()):
        alt_names = ", ".join(f"{a.get('name','?')}-{a.get('realmSlug','?')}" for a in alts)
        findings.append(Finding(
            category="Main left, alts orphaned",
            character=main_name.title(),
            realm="(left)",
            detail=f"{len(alts)} alt(s) need promotion or note rewrite: {alt_names}",
            severity="action",
        ))

    # CHECK 2: Officer-note alt syntax.
    for m in in_game:
        note = (m.get("officerNote") or "").strip()
        if note and "alt" in note.lower() and not ALT_NOTE_RE.match(note):
            findings.append(Finding(
                category="Broken alt-note syntax",
                character=m.get("name", "?"),
                realm=m.get("realmSlug", "?"),
                detail=f'Note: "{note}"  →  expected "<MainName>\'s Alt"',
                severity="warn",
            ))

    # CHECK 3-6: Discord cross-checks (only if a Discord roster was supplied).
    if discord_members:
        # Index Discord by (char, realm) and by char-only fallback.
        discord_by_char_realm: dict[tuple[str, str], DiscordMember] = {}
        discord_by_char: dict[str, list[DiscordMember]] = {}
        for d in discord_members:
            if d.char_name:
                if d.realm_slug:
                    discord_by_char_realm[(d.char_name, d.realm_slug)] = d
                discord_by_char.setdefault(d.char_name, []).append(d)

        # CHECK 3: Discord member with main role but their (char, realm) isn't in
        # any captured guild. Severity depends on whether we actually scanned
        # that realm — if we didn't capture the realm's guild, we can't honestly
        # call this an action item, so we bucket those into one INFO line.
        uncaptured_misses: dict[str, list[DiscordMember]] = {}  # realm → [d, ...]
        for d in discord_members:
            if not d.is_main:
                continue
            if not d.char_name:
                findings.append(Finding(
                    category="Discord nick missing character format",
                    character=d.nickname or d.user,
                    realm="(Discord)",
                    detail=f"User {d.user} has main role '{d.main_role}' but nickname doesn't parse as Charname-Realm.",
                    severity="warn",
                ))
                continue
            in_game_match = by_char_realm.get((d.char_name, d.realm_slug))
            if not in_game_match:
                # Fallback: maybe the realm differs (xfer); try char-only match
                if d.char_name in main_index:
                    actual = main_index[d.char_name]
                    findings.append(Finding(
                        category="Realm mismatch (Discord nick vs in-game)",
                        character=d.char_name.title(),
                        realm=d.realm_slug or "?",
                        detail=f"Discord shows {d.realm_slug or '?'} but in-game character is on {actual.get('realmSlug','?')} — update Discord nickname.",
                        severity="action",
                    ))
                elif d.realm_slug and d.realm_slug in captured_realms:
                    # We DID scan this realm — they should be in the roster, but aren't.
                    findings.append(Finding(
                        category=f"Discord {d.main_role} not in any captured guild",
                        character=d.char_name.title(),
                        realm=d.realm_slug,
                        detail=f"User {d.user} has '{d.main_role}' role but isn't in the captured guild on {d.realm_slug}. Demote to MDGA Friend or remove.",
                        severity="action",
                    ))
                else:
                    # Realm wasn't captured — bucket and report as one INFO line.
                    bucket_key = d.realm_slug or "(no realm in nick)"
                    uncaptured_misses.setdefault(bucket_key, []).append(d)

        if uncaptured_misses:
            total = sum(len(v) for v in uncaptured_misses.values())
            captured_list = ", ".join(sorted(captured_realms)) or "none"
            top_realms = sorted(uncaptured_misses.items(), key=lambda kv: -len(kv[1]))[:8]
            sample = ", ".join(f"{realm} ({len(ds)})" for realm, ds in top_realms)
            findings.append(Finding(
                category="Discord members on realms we didn't scan",
                character="-",
                realm="-",
                detail=(
                    f"{total} Discord members hold a main role but their nick's realm wasn't in your "
                    f"capture set. Captured realms: {captured_list}. Top uncaptured realms by count: "
                    f"{sample}. Re-run with /reload in each federation guild to convert these to ACTION items."
                ),
                severity="info",
            ))

        # CHECK 4: In-game main rank but no Discord presence at all.
        for char_lower, entry in main_index.items():
            if char_lower in discord_by_char:
                continue
            findings.append(Finding(
                category="In-game main not in Discord",
                character=entry.get("name", "?"),
                realm=entry.get("realmSlug", "?"),
                detail=f"Rank: {entry.get('rankName','?')} (guild {entry.get('_guild','?')}) — consider inviting to Discord or demoting in-game.",
                severity="action",
            ))

        # CHECK 5/6 (consolidated): Rank match. The user's in-game rank should be
        # represented by at least one of their Discord rank-tier roles. Officers
        # legitimately hold multiple ceremonial roles on top (Warchief +
        # War Council + RWC), so we never flag extras — only the *absence* of
        # their actual in-game rank from the Discord role set. The set of
        # rank-tier roles includes main-faction roles (Durotarian / Elwynnian),
        # since those are simultaneously the "I'm a member" badge AND a literal
        # in-game rank name. Severity is ACTION when the in-game rank is
        # ALT/TRIAL (real demotion not yet reflected in Discord), WARN otherwise.
        rank_role_universe = GAME_RANK_ROLES | MAIN_ROLES
        for d in discord_members:
            if not d.char_name:
                continue
            entry = by_char_realm.get((d.char_name, d.realm_slug)) or main_index.get(d.char_name)
            if not entry:
                continue
            in_game_rank = (entry.get("rankName") or "").strip()
            if not in_game_rank:
                continue
            held_rank_roles = d.roles & rank_role_universe
            if not held_rank_roles:
                continue  # User holds no rank-mirror roles — can't audit
            normalized_in_game = "DUROTARI" if in_game_rank.upper() == "THE DUROTARI" else in_game_rank.upper()
            held_upper = {r.upper() for r in held_rank_roles}
            if normalized_in_game not in held_upper:
                is_demotion = in_game_rank.upper() in {"ALT", "TRIAL"}
                findings.append(Finding(
                    category="Discord rank doesn't match in-game" + (" (demoted)" if is_demotion else ""),
                    character=entry.get("name", "?"),
                    realm=entry.get("realmSlug", "?"),
                    detail=(
                        f"In-game rank '{in_game_rank}' but Discord roles are: "
                        f"{', '.join(sorted(held_rank_roles))}. "
                        + ("Demote in Discord — they're no longer a main." if is_demotion
                           else "Promote in-game OR adjust Discord role.")
                    ),
                    severity="action" if is_demotion else "warn",
                ))

    # CHECK 7: Website report counts (if available).
    if website:
        for err in website.get("errors") or []:
            findings.append(Finding(
                category="Website fetch error",
                character="-", realm="-", detail=err, severity="info",
            ))
        for key in ("guild_gaps", "discord_orphans", "spelling_mismatches"):
            n = len(website.get(key) or [])
            if n:
                findings.append(Finding(
                    category=f"Website report: {key.replace('_', ' ')}",
                    character="-", realm="-",
                    detail=f"{n} row(s) — see {key.replace('_', '-')} on the admin dashboard for details.",
                    severity="info",
                ))

    return findings


# ──────────────────────────────────────────────────────────────────
# Report writers
# ──────────────────────────────────────────────────────────────────


def print_console_report(findings: list[Finding]) -> None:
    by_cat: dict[str, list[Finding]] = {}
    for f in findings:
        by_cat.setdefault(f.category, []).append(f)

    print("")
    print("================================================")
    print("  AUDIT REPORT")
    print("================================================")
    if not by_cat:
        print("  No findings — Discord and in-game rosters are in sync.")
        return
    for cat, items in sorted(by_cat.items()):
        sev = items[0].severity.upper()
        print(f"\n  [{sev}] {cat} ({len(items)})")
        for f in items[:25]:
            label = f"{f.character}-{f.realm}" if f.realm not in ("-", "(Discord)", "(left)") else f.character
            print(f"    - {label}: {f.detail}")
        if len(items) > 25:
            print(f"    ... and {len(items) - 25} more (see CSV)")


SEVERITY_FILLS = {
    "action": PatternFill(start_color="FFF8D7DA", end_color="FFF8D7DA", fill_type="solid"),  # red-ish
    "warn":   PatternFill(start_color="FFFFF3CD", end_color="FFFFF3CD", fill_type="solid"),  # amber
    "info":   PatternFill(start_color="FFD1ECF1", end_color="FFD1ECF1", fill_type="solid"),  # blue-ish
}
SEVERITY_FONT = {
    "action": Font(bold=True, color="FF842029"),
    "warn":   Font(color="FF664D03"),
    "info":   Font(color="FF055160"),
}


def write_xlsx_report(
    findings: list[Finding],
    path: str,
    capture_summary: list[str] | None = None,
    generated_by: str | None = None,
) -> None:
    """Write a formatted Excel workbook with two sheets:
       1. Summary — count of findings per category, plus the capture summary.
       2. Findings — every finding as a row in an Excel Table (filterable, banded).
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = "MDGA Audit Report"
    summary_ws["A1"].font = Font(bold=True, size=16)
    summary_ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_ws["A2"].font = Font(italic=True, color="FF666666")
    row = 3
    if generated_by:
        summary_ws.cell(row=row, column=1, value=f"Run by: {generated_by}").font = Font(italic=True, color="FF666666")
        row += 1
    row += 1  # blank spacer
    if capture_summary:
        summary_ws.cell(row=row, column=1, value="Captured guilds").font = Font(bold=True)
        row += 1
        for line in capture_summary:
            summary_ws.cell(row=row, column=1, value=line)
            row += 1
        row += 1

    summary_ws.cell(row=row, column=1, value="Findings by category").font = Font(bold=True)
    row += 1
    summary_ws.cell(row=row, column=1, value="Category").font = Font(bold=True)
    summary_ws.cell(row=row, column=2, value="Count").font = Font(bold=True)
    summary_ws.cell(row=row, column=3, value="Highest severity").font = Font(bold=True)
    row += 1

    by_cat: dict[str, list[Finding]] = {}
    for f in findings:
        by_cat.setdefault(f.category, []).append(f)
    severity_rank = {"action": 3, "warn": 2, "info": 1}
    for cat, items in sorted(by_cat.items(), key=lambda kv: (-max(severity_rank.get(i.severity, 0) for i in kv[1]), kv[0])):
        sev = max(items, key=lambda i: severity_rank.get(i.severity, 0)).severity
        summary_ws.cell(row=row, column=1, value=cat)
        summary_ws.cell(row=row, column=2, value=len(items))
        sev_cell = summary_ws.cell(row=row, column=3, value=sev.upper())
        sev_cell.fill = SEVERITY_FILLS.get(sev, PatternFill())
        sev_cell.font = SEVERITY_FONT.get(sev, Font())
        row += 1

    summary_ws.column_dimensions["A"].width = 56
    summary_ws.column_dimensions["B"].width = 10
    summary_ws.column_dimensions["C"].width = 18

    # ── Sheet 2: Findings (Excel Table — filterable + banded) ──
    findings_ws = wb.create_sheet("Findings")
    headers = ["Category", "Severity", "Character", "Realm", "Detail"]
    findings_ws.append(headers)
    for f in findings:
        findings_ws.append([f.category, f.severity.upper(), f.character, f.realm, f.detail])
    # Style header
    for col_idx, _ in enumerate(headers, 1):
        cell = findings_ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
    # Color the Severity column per row
    for row_idx in range(2, len(findings) + 2):
        sev = (findings_ws.cell(row=row_idx, column=2).value or "").lower()
        cell = findings_ws.cell(row=row_idx, column=2)
        cell.fill = SEVERITY_FILLS.get(sev, PatternFill())
        cell.font = SEVERITY_FONT.get(sev, Font())

    # Wrap the Detail column; left-align everything; size columns sensibly
    widths = {1: 38, 2: 12, 3: 22, 4: 18, 5: 80}
    for idx, w in widths.items():
        findings_ws.column_dimensions[get_column_letter(idx)].width = w
    for row_idx in range(2, len(findings) + 2):
        findings_ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    findings_ws.row_dimensions[1].height = 22
    findings_ws.freeze_panes = "A2"

    # Wrap the data range in an Excel Table so filters/banding are native.
    if findings:
        last_col = get_column_letter(len(headers))
        last_row = len(findings) + 1
        table = Table(displayName="AuditFindings", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        findings_ws.add_table(table)
    else:
        findings_ws.cell(row=2, column=1, value="No findings — Discord and in-game rosters are in sync.")

    wb.save(path)


# Kept for callers that explicitly want CSV; xlsx is now the default.
def write_csv_report(findings: list[Finding], path: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Category", "Severity", "Character", "Realm", "Detail"])
        writer.writeheader()
        for finding in findings:
            writer.writerow(finding.as_row())


# ──────────────────────────────────────────────────────────────────
# Main flow
# ──────────────────────────────────────────────────────────────────


def prompt(label: str, default: str | None = None) -> str:
    suffix = f" [{default}]" if default else ""
    while True:
        ans = input(f"  {label}{suffix}: ").strip()
        if ans:
            return ans
        if default is not None:
            return default


def prompt_int(label: str, default: int) -> int:
    while True:
        ans = input(f"  {label} [{default}]: ").strip()
        if not ans:
            return default
        try:
            return int(ans)
        except ValueError:
            print("    Please enter a number.")


def prompt_yes_no(label: str, default: bool = True) -> bool:
    suffix = "[Y/n]" if default else "[y/N]"
    ans = input(f"  {label} {suffix}: ").strip().lower()
    if not ans:
        return default
    return ans in ("y", "yes")


def pick_account(wow_path: str) -> str | None:
    accounts = list_account_folders(wow_path)
    if not accounts:
        print(f"  ! No accounts in {os.path.join(wow_path, 'WTF', 'Account')}")
        return prompt("Type the account folder name manually") or None
    if len(accounts) == 1:
        print(f"  Using account: {accounts[0]}")
        return accounts[0]
    print("  Multiple WoW accounts found:")
    for i, a in enumerate(accounts, 1):
        print(f"    {i}) {a}")
    while True:
        ans = input(f"  Pick [1-{len(accounts)}]: ").strip()
        try:
            n = int(ans)
            if 1 <= n <= len(accounts):
                return accounts[n - 1]
        except ValueError:
            pass
        print("    Invalid selection.")


def main() -> int:
    print("")
    print("================================================")
    print(f"  MDGA Audit Tool v{VERSION}")
    print("================================================")
    print("")

    # Make sure discord-reports/ + archive/ exist next to the exe.
    setup_workdir()

    # Load saved config so prompts pre-fill with last-used values. --reconfigure
    # forces a clean wizard; otherwise just press Enter at each prompt to accept.
    force_reconfigure = "--reconfigure" in sys.argv or "--setup" in sys.argv
    config = {} if force_reconfigure else load_config()
    if config and not force_reconfigure:
        print(f"  Loaded config from {os.path.basename(CONFIG_FILE)} — press Enter at any prompt to use the saved value.")
        print(f"  (Run with --reconfigure to wipe and start fresh.)")
        print("")

    # 1) Locate WoW + account
    wow_path = config.get("wow_path") or detect_wow_path()
    if wow_path:
        print(f"  Detected WoW: {wow_path}")
        if not prompt_yes_no("Use this path?", True):
            wow_path = None
    while not wow_path:
        p = prompt("WoW _retail_ folder path").strip('"')
        if os.path.isdir(p):
            wow_path = p
        else:
            print(f"    Path not found: {p}")

    account = config.get("account_name")
    if account:
        if os.path.isdir(os.path.join(wow_path, "WTF", "Account", account)):
            print(f"  Using saved account: {account}")
        else:
            print(f"  Saved account '{account}' not found in WTF/Account — re-picking.")
            account = None
    if not account:
        account = pick_account(wow_path)
    if not account:
        print("  ! Cannot continue without account.")
        return 1
    sv_path = os.path.join(wow_path, "WTF", "Account", account, "SavedVariables", "MDGA.lua")
    if not os.path.exists(sv_path):
        print(f"  ! No MDGA.lua at {sv_path}")
        print("  ! Make sure the MDGA addon is installed and you've /reloaded at least once.")
        return 1

    # 2) Capture N guilds
    n_guilds = prompt_int("How many guilds will you capture", config.get("n_guilds", 1))
    snapshots: list[GuildSnapshot] = []
    baseline_mtime = os.path.getmtime(sv_path)
    for i in range(1, n_guilds + 1):
        print("")
        print(f"  ── Guild #{i} of {n_guilds} ──")
        print(f"    Log into a character in guild #{i}, click 'Generate Report' in")
        print(f"    the MDGA addon, then click Reload. I'll detect it.")
        snap = watch_for_capture(sv_path, baseline_mtime)
        if not snap:
            print("  ! Timeout waiting for /reload. Aborting.")
            return 1
        print(f"  ✓ Captured: {snap.guild_name} ({len(snap.members)} members)")
        snapshots.append(snap)
        baseline_mtime = os.path.getmtime(sv_path)

    # 3) Discord roster — fixed location: <exe>/discord-reports/. If empty, bail
    #    with explicit instructions so the officer knows where to drop the bot
    #    CSVs. Files are moved into the run's archive after a successful audit.
    discord_csvs = sorted(
        os.path.join(DISCORD_INPUT_DIR, f) for f in os.listdir(DISCORD_INPUT_DIR)
        if f.lower().endswith(".csv")
    ) if os.path.isdir(DISCORD_INPUT_DIR) else []
    if not discord_csvs:
        print("")
        print(f"  ! No Discord CSVs found in: {DISCORD_INPUT_DIR}")
        print( "  ! Steps to fix:")
        print( "  !   1. Open Discord → #bot-spam channel")
        print( "  !   2. Download the latest two roster CSV exports (Durotarian + Elwynnian)")
        print(f"  !   3. Drop both files into:  {DISCORD_INPUT_DIR}")
        print( "  !   4. Re-run this tool")
        return 1
    print("")
    print(f"  Found {len(discord_csvs)} Discord CSV(s) in {DISCORD_INPUT_DIR}:")
    for c in discord_csvs:
        print(f"    - {os.path.basename(c)}")
    discord_members = read_discord_csvs(discord_csvs)
    main_count = sum(1 for m in discord_members if m.is_main)
    print(f"  ✓ Loaded {len(discord_members)} Discord members ({main_count} with main rank).")

    # 4) Website pull (optional). The token is also used to attribute the report.
    website: dict[str, Any] | None = None
    generated_by: str | None = None
    server: str = config.get("server_url") or DEFAULT_SERVER
    token: str = config.get("token") or ""
    if prompt_yes_no("Pull current website roster from mdga.gg?", True):
        server = prompt("Server URL", server)
        token = prompt("Token (paste from /profile)", default=token if token else None) if not token else (
            prompt("Token (Enter to use saved)", default=token)
        )
        payload = decode_jwt_payload(token)
        username = payload.get("username") or "unknown"
        rank = payload.get("rank") or ""
        generated_by = f"{username}{f' ({rank})' if rank else ''}"
        try:
            website = fetch_website_audit_data(server, token)
            print(f"  ✓ Fetched website reports")
        except Exception as e:
            print(f"  ! Website fetch failed: {e}")
            website = None

    # 5) Audit
    findings = audit(snapshots, discord_members, website)
    print_console_report(findings)

    # 6) Output: create archive/<timestamp>/ holding xlsx + moved inputs.
    run_dir = make_run_dir()
    xlsx_path = os.path.join(run_dir, "MDGA_audit.xlsx")
    capture_summary = [f"{snap.guild_name} ({len(snap.members)} members)" for snap in snapshots]
    write_xlsx_report(findings, xlsx_path, capture_summary=capture_summary, generated_by=generated_by)

    # 6a) Snapshot WoW rosters as JSON for re-audit / longitudinal compare.
    for snap in snapshots:
        archive_snapshot_json(snap, os.path.join(run_dir, "wow"))

    # 6b) Move the Discord CSVs from discord-reports/ into the run archive so
    #     the input folder is clean for the next run.
    import shutil
    for csv_path in discord_csvs:
        try:
            shutil.move(csv_path, os.path.join(run_dir, "discord", os.path.basename(csv_path)))
        except Exception as e:
            print(f"  ! Couldn't move {os.path.basename(csv_path)} into archive: {e}")

    # 7) Persist config so the next run skips most prompts.
    save_config({
        "wow_path": wow_path,
        "account_name": account,
        "server_url": server,
        "token": token,
        "n_guilds": n_guilds,
    })

    print(f"\n  Run archived → {run_dir}")
    print(f"  Excel report  → {xlsx_path}")
    if generated_by:
        print(f"  Report attributed to: {generated_by}")
    print( "  discord-reports/ is now empty — drop fresh CSVs there for the next run.")
    print("")
    return 0


if __name__ == "__main__":
    try:
        rc = main()
    except KeyboardInterrupt:
        print("\n  Cancelled.")
        rc = 130
    except Exception as e:
        print(f"\n  ! Fatal: {e}")
        import traceback
        traceback.print_exc()
        rc = 1
    # When run by double-click, hold the window open so the user sees output.
    if sys.stdin and sys.stdin.isatty():
        input("\n  Press Enter to close.")
    sys.exit(rc)
